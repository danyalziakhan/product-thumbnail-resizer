from __future__ import annotations

import os
import shutil
import sys

from concurrent.futures import ThreadPoolExecutor
from contextlib import suppress
from datetime import datetime
from pathlib import Path
from threading import Lock
from time import sleep
from typing import TYPE_CHECKING

import cv2
import cv2_ext
import numpy as np
import openpyxl
import pandas as pd
import PIL.Image
import requests

from openpyxl.styles import PatternFill
from requests_cache import CachedSession
from tqdm import tqdm

from product_thumbnail_resizer.log import logger
from product_thumbnail_resizer.settings import Settings


if TYPE_CHECKING:
    from typing import Any, Final

SCRIPT_PATH: Final[str] = os.path.dirname(os.path.realpath(sys.argv[0]))
LOG_DIR = os.path.join(SCRIPT_PATH, "logs")
DOWNLOAD_IMAGES_DIR: Final[str] = "downloaded_images"
RESIZED_IMAGES_DIR: Final[str] = "resized_images"

TODAY_DATE = f"{datetime.now().strftime('%Y%m%d')}"

MAX_WORKERS: Final[int] = 150

lock: Lock = Lock()


def format_excel_openpyxl(filename: str):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):  # type: ignore
        for cell in rows:
            cell.fill = PatternFill(
                start_color=f"{153:02x}{51:02x}{0:02x}",
                end_color=f"{153:02x}{51:02x}{0:02x}",
                fill_type="solid",
            )

    wb.save(filename)


def download_images(
    thumbnail: str, brand_name: str, model_name: str, bad_links: list[str]
):
    thumbnail = thumbnail.replace("\\", "/")

    save_path_bad_link = os.path.join(
        DOWNLOAD_IMAGES_DIR, brand_name, f"{model_name}_bad_link.jpg"
    )

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36"
    }

    session = CachedSession("images_cache")

    try:
        response: requests.Response = session.get(
            thumbnail, stream=True, headers=headers
        )

        if response.status_code == 200:
            response.raw.decode_content = True
            save_valid_image(brand_name, model_name, response.raw)

        # ? 403 means "Forbidden"
        # ? In our case, it happens when there are mix of backward and forward slashes in the URLs
        elif response.status_code == 403:
            # ? Use requests object directly and not the cached session (we don't want to fetch invalid cached 403 response again!)
            response: requests.Response = requests.get(
                thumbnail, stream=True, headers=headers
            )

            if response.status_code == 200:
                response.raw.decode_content = True
                save_valid_image(brand_name, model_name, response.raw)
            else:
                save_invalid_image(thumbnail, bad_links, save_path_bad_link, response)
        else:
            save_invalid_image(thumbnail, bad_links, save_path_bad_link, response)

    except ConnectionResetError:
        logger.warning(f"Connection Error: {thumbnail}")
        sleep(1)
        download_images(thumbnail, brand_name, model_name, bad_links)
    except Exception as e:
        logger.error(e)
        raise ValueError(f"Couldn't download the image {thumbnail}") from e


def save_valid_image(brand_name: str, model_name: str, raw_data: Any):
    with open(
        os.path.join(DOWNLOAD_IMAGES_DIR, brand_name, f"{model_name}.jpg"), "wb"
    ) as f:
        shutil.copyfileobj(raw_data, f)


def save_invalid_image(
    thumbnail: str, bad_links: list[str], save_path_bad_link: str, r: requests.Response
):
    logger.warning(
        f"URL ({thumbnail}) doesn't contain the image <blue>(Status code: {r.status_code})</>"
    )
    bad_links.append(thumbnail)
    with open(save_path_bad_link, "wb"):
        pass


def resize_images(model_name: str, product_thumbnail: str, brand_name: str):
    image_path = os.path.join(DOWNLOAD_IMAGES_DIR, brand_name, f"{model_name}.jpg")
    save_path = os.path.join(RESIZED_IMAGES_DIR, brand_name, f"{model_name}.jpg")
    save_path_checked = os.path.join(
        RESIZED_IMAGES_DIR, brand_name, f"{model_name}_checked.jpg"
    )
    with lock:
        # ? If image was invalid or was already resized, it won't be present in the directory
        try:
            img: np.ndarray[Any, Any] = cv2_ext.imread(image_path)  # type: ignore
        except FileNotFoundError:
            return
    try:
        if img.shape[0] <= 600:
            img = cv2.resize(img, (1000, 1000))  # type: ignore
            with lock:
                cv2_ext.imwrite(save_path, img)
        elif not (product_thumbnail.endswith(".jpg")):
            with lock:
                cv2_ext.imwrite(save_path, img)
        else:
            logger.debug(
                f"{model_name}.jpg is already greater than 600: {img.shape[0]}x{img.shape[1]}, so we are not resizing it."
            )
            with lock:
                shutil.copy(image_path, save_path_checked)
    except AttributeError:
        # https://stackoverflow.com/questions/21669657/getting-cannot-write-mode-p-as-jpeg-while-operating-on-jpg-image

        # ? Replacing transparent background to white background for saving in JPEG
        # ? See: https://stackoverflow.com/questions/50898034/how-replace-transparent-with-a-color-in-pillow
        img_pillow = PIL.Image.open(image_path).convert("RGBA")  # type: ignore

        new_image = PIL.Image.new(
            "RGBA",
            img_pillow.size,  # type: ignore
            "WHITE",  # type: ignore
        )  # Create a white rgba background
        new_image.paste(
            img_pillow,  # type: ignore
            (0, 0),
            img_pillow,  # type: ignore
        )  # Paste the image on the background
        img_pillow = new_image.convert("RGB")  # type: ignore

        if img_pillow.width <= 600 or img_pillow.height <= 600:
            img_pillow = img_pillow.resize((1000, 1000))  # type: ignore
            with lock:
                img_pillow.save(save_path, "JPEG")
        elif not (product_thumbnail.endswith(".jpg")):
            with lock:
                img_pillow.save(save_path, "JPEG")
        else:
            logger.debug(
                f"{model_name}.jpg is already greater than 600: {img_pillow.width}x{img_pillow.height}, so we are not resizing it."
            )
            with lock:
                shutil.copy(image_path, save_path_checked)


def get_url(model_name: str, brand_name: str):
    checked_image_path = os.path.join(
        RESIZED_IMAGES_DIR, brand_name, f"{model_name}_checked.jpg"
    )
    bad_link_image_path_ = os.path.join(
        DOWNLOAD_IMAGES_DIR, brand_name, f"{model_name}_bad_link.jpg"
    )
    if os.path.exists(checked_image_path):
        with suppress(OSError):
            os.remove(checked_image_path)
        return "checked"
    elif os.path.exists(bad_link_image_path_):
        with suppress(OSError):
            os.remove(bad_link_image_path_)
        return "NOT PRESENT"
    else:
        image_path = os.path.join(RESIZED_IMAGES_DIR, brand_name, f"{model_name}.jpg")
        filename = Path(image_path).name
        return f"http://gi.esmplus.com/lala01281/{brand_name}/{filename}"


def run(settings: Settings) -> None:
    logger.log("ACTION", f"Reading: {settings.input_file} ...")

    dataframe = parse_excel_sheet(
        settings.input_file, settings.sheet_name, settings.model_name_column
    )

    try:
        dataframe[settings.thumbnail_column]
    except KeyError as e:
        error = f'"{settings.thumbnail_column}" column is not present in file "{os.path.basename(settings.input_file)}"'
        raise KeyError(error) from e

    try:
        dataframe[settings.model_name_column]
    except KeyError as e:
        error = f'"{settings.model_name_column}" column is not present in file "{os.path.basename(settings.input_file)}"'
        raise KeyError(error) from e

    product_thumbnails: list[str] = dataframe[settings.thumbnail_column].to_list()  # type: ignore
    model_names: list[str] = dataframe[settings.model_name_column].to_list()  # type: ignore

    logger.info(f"Brand name: {settings.brand_name}")

    os.makedirs(Path(DOWNLOAD_IMAGES_DIR, settings.brand_name), exist_ok=True)
    os.makedirs(Path(RESIZED_IMAGES_DIR, settings.brand_name), exist_ok=True)

    assert product_thumbnails, (
        "일치하는 제품이 없으므로 이미지가 다운로드되지 않습니다. 이제 프로그램이 중지됩니다."
    )

    try:
        assert len(model_names) == len(product_thumbnails)
    except AssertionError as e:
        error = f"Model names and Product Thumbnails list are not the same length: {len(model_names)} vs {len(product_thumbnails)}"
        raise AssertionError(error) from e

    logger.log("ACTION", "Downloading images ...")

    bad_links: list[str] = []

    # ? Don't exceed the MAX_WORKERS count
    max_workers = min(len(product_thumbnails), MAX_WORKERS)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        iterator: Any = tqdm(zip(product_thumbnails, model_names))
        for idx, (thumbnail, model_name) in enumerate(iterator):
            try:
                thumbnail = thumbnail.strip()
            except AttributeError as err:
                raise ValueError(f"Thumbnail is empty for row # {idx + 2}") from err
            executor.submit(
                download_images,
                thumbnail,
                settings.brand_name,
                model_name.strip(),
                bad_links,
            )

    logger.log("ACTION", "Resizing images ...")
    save_path = resize(
        dataframe,
        settings.brand_name,
        settings.thumbnail_new_column,
        product_thumbnails,
        model_names,
        bad_links,
    )

    # ? Final assertions
    df = pd.read_excel(save_path)
    df = df[df[settings.thumbnail_new_column].notna()]
    df = df[~df[settings.thumbnail_new_column].str.contains("checked")]

    total_resized_images = len(
        os.listdir(os.path.join(RESIZED_IMAGES_DIR, settings.brand_name))
    )
    total_new_thumbnail_urls = len(df[settings.thumbnail_new_column].to_list())  # type: ignore
    assert total_resized_images == total_new_thumbnail_urls, (
        f"The number of images in 'resized_images' directory ({total_resized_images}) is not the same as number of new thumbnail URLs ({total_new_thumbnail_urls})"
    )  # type: ignore


def resize(
    dataframe: pd.DataFrame,
    brand_name: str,
    thumbnail_new_column: str,
    product_thumbnails: list[str],
    model_names: list[str],
    bad_links: list[str],
):
    iterator: Any = tqdm(zip(model_names, product_thumbnails))
    for model_name, product_thumbnail in iterator:
        resize_images(model_name, product_thumbnail, brand_name)

    logger.log("ACTION", "Saving images to thumbnail new column ...")
    prd_list: list[str] = [
        get_url(model_name, brand_name) for model_name in model_names
    ]

    assert prd_list, "Product Thumbnail List is empty"
    assert len(prd_list) == len(product_thumbnails), "Rows are not equal"

    # ValueError
    dataframe[thumbnail_new_column].iloc[:] = prd_list  # type: ignore

    save_path = os.path.join(SCRIPT_PATH, "RESULT.xlsx")

    if os.path.exists(save_path):
        os.remove(save_path)

    dataframe.to_excel(save_path, engine="openpyxl", index=False)

    format_excel_openpyxl(save_path)

    assert len(product_thumbnails) == len(model_names), "Column lengths are not equal"

    assert len(os.listdir(Path(DOWNLOAD_IMAGES_DIR))) != 0, (
        "Downloaded directory is empty"
    )
    assert len(os.listdir(Path(DOWNLOAD_IMAGES_DIR, brand_name))) != 0, (
        "Downloaded directory is empty"
    )

    files_present: list[str] = []
    for brand in model_names:
        if str(brand) != "nan":
            path = os.path.join(
                Path(DOWNLOAD_IMAGES_DIR, brand_name), f"{str(brand)}.jpg"
            )
            if not os.path.exists(path):
                continue

            files_present.append(brand)

    assert len(os.listdir(Path(DOWNLOAD_IMAGES_DIR, brand_name))) == len(
        set(files_present)
    ), (
        f"Downloaded images are not equal as number of rows: {len(os.listdir(Path(DOWNLOAD_IMAGES_DIR, brand_name)))} vs {len(set(files_present))}"
    )

    all_resized_images = os.listdir(Path(RESIZED_IMAGES_DIR, brand_name))
    assert len(os.listdir(Path(RESIZED_IMAGES_DIR))) != 0, "Resized directory is empty"
    assert len(all_resized_images) != 0, "Resized directory is empty"

    with open("bad_links.txt", mode="w") as file:
        for link in bad_links:
            file.write(link + "\n")

    if bad_links:
        logger.warning(f"There are total {len(bad_links)} bad links")

    return save_path


def parse_excel_sheet(
    input_file: str, sheet_name: str, model_name_column: str
) -> pd.DataFrame:
    """
    Reads an Excel file into a pandas DataFrame.

    Args:
        input_file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet in the Excel file.
        model_name_column (str): The column containing model names.

    Returns:
        pd.DataFrame: A pandas DataFrame containing the data from the specified sheet.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the sheet name is invalid or if the model_name_column contains non-model-related data.
        Exception: Any other errors that may occur while processing the data file.
    """
    try:
        # Try to read the Excel file
        excel_file = pd.ExcelFile(
            os.path.join(SCRIPT_PATH, input_file), engine="openpyxl"
        )

        if not excel_file.sheet_names or sheet_name not in excel_file.sheet_names:
            raise ValueError(
                f"The sheet '{sheet_name}' does not exist in the Excel file."
            )

        # Read the specified sheet
        df = excel_file.parse(sheet_name)
    except FileNotFoundError as e:
        logger.error(f"The input file '{input_file}' was not found. {e}")
        raise

    try:
        df.dropna(how="all").dropna(how="any", subset=[model_name_column])
    except Exception as e:
        logger.error(f"An error occurred while processing the data file. {e}")
        raise

    return df
